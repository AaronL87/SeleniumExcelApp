from selenium import webdriver
import openpyxl
from datetime import datetime
from math import ceil

class excelApp:
    def __init__(self):
        self.options = webdriver.ChromeOptions()

        # Hides browser
        ####self.options.add_argument('headless')

        # Optional code for downloading PDFs:
        # self.options.add_experimental_option('prefs', {
        # "download.default_directory": r'C:\Users\Aaron\Desktop\VBA finance advisor app/pdf downloads', #Change default directory for downloads
        # "download.prompt_for_download": False, #To auto download the file
        # "download.directory_upgrade": True,
        # "plugins.always_open_pdf_externally": True #It will not show PDF directly in chrome
        # })

        self.linkToExcel()

        self.getBrowser()

        self.page = 1
        self.url = 'https://dash.lead.ac/repositories?page='+str(self.page)

        self.goToWebsite()
        
        # If not logged in:
        if self.driver.current_url == 'https://dash.lead.ac/users/sign_in':
            self.tryLoggingIn()

        # If still not logged in:
        if self.driver.current_url != self.url:
            self.loginError()

        self.getLastUpdate()
        
        # Counter determines whether Excel sheet is already up-to-date
        # If counter==0 when dates match, it is
        # If counter>0 when dates match, it isn't and stores new date
        self.counter = 0

        self.excelDict = {}
        self.createKeysForExcelDictionary()

        self.updateTables()
        
        #Failsafe
        self.saveAndExit()

    def getBrowser(self):
        try:
            self.driver = webdriver.Chrome(executable_path=r'C:\Users\Aaron\Desktop\VBA finance advisor app/chromedriver',
                                        options=self.options)
        except:
            self.firstSheet['D1'] = 'Webdriver Error'
            self.saveAndExit()

    def linkToExcel(self):
        self.wb = openpyxl.load_workbook(r'C:\Users\Aaron\Desktop\VBA finance advisor app\ExcelFile.xlsx')
        self.sheetNames = self.wb.sheetnames
        self.firstSheet = self.wb[self.sheetNames[0]]
        self.firstSheet['D1'],self.firstSheet['F1'] = None,None # Resets 'error' cells

    def goToWebsite(self):
        self.driver.get(self.url)

    def tryLoggingIn(self):
        username = self.driver.find_element_by_id('user_email')
        password = self.driver.find_element_by_id('user_password')
        button = self.driver.find_element_by_name('button')

        username.send_keys('username')
        password.send_keys('password')
        button.click()

    def loginError(self):
        self.firstSheet['D1'] = 'Login Error'
        self.saveAndExit()

    def saveAndExit(self):
        self.driver.close()
        
        try:
            self.wb.save(r'C:\Users\Aaron\Desktop\VBA finance advisor app\ExcelFile.xlsx')
        except:
            pass # If spreadsheet is open while running app
        exit()
    
    def createKeysForExcelDictionary(self):
        for name in self.sheetNames:
            self.excelDict.update({name:None})

    def getLastUpdate(self):
        if type(self.firstSheet['B1'].value)==datetime:
            self.lastUpdated = self.firstSheet['B1'].value
        elif type(self.firstSheet['B1'].value)==str:
            self.lastUpdated = datetime.strptime(self.firstSheet['B1'].value,'%Y-%m-%d %H:%M%p')
        else:
            self.firstSheet['D1'] = 'Last Updated cannot be of the type {}'.format(type(self.firstSheet['B1'].value))
            self.saveAndExit()

    def updateTables(self):
        self.findRowTotal()
        
        for page in self.totalPages:
            if self.page != 1:
                self.goToWebsite()
            
            self.webTable = self.driver.find_element_by_id('triage_form')
            self.webRows = self.webTable.find_elements_by_tag_name('tr')

            self.scrapeWebTableAndUpdateExcel()

            self.page+=1

    def findRowTotal(self):
        firstWebTable = self.driver.find_element_by_id('triage_form')
        rowTotalFooter = firstWebTable.find_element_by_class_name('dataTables_info').text
        self.totalPages = ceil(int(rowTotalFooter.split()[-2])/25) # Takes total rows and divides by number of rows per page (25)

    def scrapeWebTableAndUpdateExcel(self):
        for webRow in self.webRows[1:]:
            self.tempRowData = []
            for col in webRow.find_elements_by_tag_name('td'):
                self.tempRowData.append(col.text)
            
            self.Returned_Date = datetime.strptime(self.tempRowData[7],'%Y-%m-%d %H:%M%p')
            if self.Returned_Date > self.lastUpdated: # Only executes if Excel needs updating
                if self.counter == 0:
                    self.newLastUpdated = self.Returned_Date

                self.currentSheetKey = self.tempRowData[8]
                self.checkForSheetName()

                if self.excelDict[self.currentSheetKey] == None:
                    self.createValuesForExcelDictionary()

                self.First_name = self.tempRowData[2]
                self.Last_name = self.tempRowData[3]            
                self.Zip_code = self.tempRowData[4] # Used in case of duplicate names
                self.PDF_link = webRow.find_element_by_link_text('PDF').get_attribute('href') #column index 12

                # Checks for customer name in sheet
                try:
                    self.rowData = self.excelDict[self.currentSheetKey][(self.First_name,self.Last_name,self.Zip_code)]
                except:
                    if self.currentSheet['F1'].value == None: # Updates 'Name Errors'
                        self.currentSheet['F1'] = self.First_name+' '+self.Last_name
                    else:
                        self.currentSheet['F1'] = self.currentSheet['F1'].value+', '+self.First_name+' '+self.Last_name
                    continue # Continue can be returned from a function, so left it in
   
                self.updateExcel()

                self.counter += 1
            else:
                if self.counter != 0:
                    self.firstSheet['B1'] = self.newLastUpdated
                self.saveAndExit()

    def checkForSheetName(self): # Checks for sheet name in workbook
        try:
            self.currentSheet = self.wb[self.currentSheetKey]
        except:
            self.firstSheet['D1'] = 'Sheet Name Error: {} does not exist.'.format(self.currentSheetKey)
            self.saveAndExit()

    def createValuesForExcelDictionary(self):
        self.checkForTableName()
        tempDict={}
        for excelRow in self.currentSheet[self.excelTable.ref][1:]:
            tempDict.update({(excelRow[3].value,excelRow[4].value,excelRow[9].value):(excelRow[1].value,excelRow[12],excelRow[0].row)})
            # Of the form {(FirstName,LastName,ZipCode):(LastUpdated,PDF_hyperlink,Row)}
        self.excelDict[self.currentSheetKey]=tempDict

    def checkForTableName(self):
        try:
            self.excelTable = self.findExcelTable('Table'+self.currentSheetKey,self.currentSheet._tables)
        except:
            self.firstSheet['D1'] = 'Table Name Error: Table{0} does not exist in sheet {0}.'.format(self.currentSheetKey)
            self.saveAndExit()

    def findExcelTable(self,tableName,tables):
        for table in tables:
            if table.displayName == tableName:
                return table

    def updateExcel(self):
        if type(self.rowData[0]) == None:
            self.currentSheet.cell(None,self.rowData[2],1).value = self.Returned_Date
        elif type(self.rowData[0]) == datetime:
            if self.rowData[0]<=self.Returned_Date: # 'equals to' in '<=' allows for corrections while rerunning following errors.
                self.currentSheet.cell(None,self.rowData[2],1).value = self.Returned_Date
                if self.rowData[1] == None:
                    self.currentSheet.cell(None,self.rowData[2],12).value = self.PDF_link
                else:
                    self.currentSheet.cell(None,self.rowData[2],12).value = self.currentSheet.cell(None,self.rowData[2],12).value+'; '+self.PDF_link
        else:
            self.firstSheet['D1'] = 'The Received Date column in row {} must be empty or a date'.format(self.rowData[2])
            self.saveAndExit()


excelApp()
